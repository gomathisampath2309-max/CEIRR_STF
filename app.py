import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from io import BytesIO

# --- PASSWORD PROTECTION ---
st.title("🧪 CEIRR Sample Collection Summary")
password = st.text_input("Enter Password:", type="password")
if password != "ceirr123":  # change password if needed
    st.warning("Please enter the correct password to access data.")
    st.stop()

# --- Load Google Sheet ---
def load_sheet(sheet_id):
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?tqx=out:csv"
    return pd.read_csv(url, on_bad_lines="skip")

# Replace this with your actual sheet ID
df = load_sheet("1cVu_jloglMcVNQUm0RHZZEKAKDcMlu4j5vpXz_KiFD4")

# --- Data Cleaning ---
df.columns = df.columns.str.strip().str.lower()
df["submissiondate"] = pd.to_datetime(df["submissiondate"], errors="coerce").dt.tz_localize(None)

# --- Filter Today’s Data ---
today_str = pd.Timestamp.today().strftime("%Y-%m-%d")
df_today = df[df["submissiondate"].dt.strftime("%Y-%m-%d") == today_str].copy()

# --- Create Sample ID ---
df_today["sample_id"] = np.where(
    df_today["sample_scan"].notna() & (df_today["sample_scan"].astype(str).str.strip() != ""),
    df_today["sample_scan"],
    df_today["sample_scan_manually"]
)

# --- Map Type Cohort and Sample Type ---
sample_type_map = {
    "1": "Nasal swab",
    "2": "Blood",
    "3": "Mucosal"
}

type_cohort_map = {
    "1": "Screening",
    "2": "Prior Infected",
    "3": "Not Infected"
}

df_today["sample_type"] = df_today["sample_type"].astype(str).map(sample_type_map).fillna(df_today["sample_type"])
df_today["type_cohort"] = df_today["type_cohort"].astype(str).map(type_cohort_map).fillna(df_today["type_cohort"])

# --- Select Final Columns ---
table = df_today[["sample_id", "sample_date_time", "type_cohort", "sample_type"]].copy()
table.columns = ["Sample ID", "Sample Date/Time", "Cohort Type", "Sample Type"]

# --- Display Table ---
st.subheader("📋 Today's Sample Collection Details")
if table.empty:
    st.warning("No sample collections found for today.")
else:
    st.dataframe(table, use_container_width=True)

    # --- Download as Excel ---
    excel_filename = f"{datetime.today().strftime('%d-%m-%Y')}_CEIRR_SampleCollection.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Today_Samples"

    # Formatting styles
    border = Border(left=Side(style="thin"), right=Side(style="thin"), 
                    top=Side(style="thin"), bottom=Side(style="thin"))
    bold_center = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center")

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=table.shape[1])
    title_cell = ws.cell(row=1, column=1, value="CEIRR Daily Sample Collection Summary")
    title_cell.font, title_cell.alignment = Font(bold=True, size=12), align_center

    # Header Row
    for col_num, column_title in enumerate(table.columns, 1):
        cell = ws.cell(row=2, column=col_num, value=column_title)
        cell.font = bold_center
        cell.alignment = align_center
        cell.border = border

    # Data Rows
    for row_num, row_data in enumerate(table.values, 3):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border

    # Save Excel to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    st.download_button(
        label="⬇️ Download Excel",
        data=buffer,
        file_name=excel_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
